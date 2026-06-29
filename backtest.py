import sqlite3
import pandas as pd
import pickle
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_FILE    = "gap_log.db"
MODEL_FILE = "model.pkl"

CONFIDENCE_THRESHOLD = 0.70
CLOSE_THRESHOLD      = 0.30   # gap_pct below this = closed
HORIZON_HOURS        = 48     # max hours we hold a trade
POSITION_SIZE        = 1000   # dollars per trade

DAY_MAP = {
    "Monday": 0, "Tuesday": 1, "Wednesday": 2,
    "Thursday": 3, "Friday": 4, "Saturday": 5, "Sunday": 6
}

def load_data():
    con = sqlite3.connect(DB_FILE)
    df  = pd.read_sql("SELECT * FROM gaps ORDER BY id", con)
    con.close()
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    return df

def split_data(df):
    # Use the last 20% of time as the holdout — data the model never trained on
    cutoff = df["timestamp"].quantile(0.80)
    train  = df[df["timestamp"] <= cutoff]
    test   = df[df["timestamp"] >  cutoff]
    print(f"  Training period: {train['timestamp'].min().strftime('%a %d %b %H:%M')} → {train['timestamp'].max().strftime('%a %d %b %H:%M')}", flush=True)
    print(f"  Test period:     {test['timestamp'].min().strftime('%a %d %b %H:%M')} → {test['timestamp'].max().strftime('%a %d %b %H:%M')}", flush=True)
    print(f"  Test rows:       {len(test):,}", flush=True)
    return test

def build_features(df):
    df = df.copy()
    df["day_num"]    = df["day_of_week"].map(DAY_MAP)
    df["is_weekend"] = df["day_num"].isin([5, 6]).astype(int)
    df["gap_abs"]    = df["gap_pct"].abs()
    return df

FEATURES = [
    "gap_abs", "is_weekend", "day_num", "hour_utc",
    "hours_to_open", "vix", "bid_ask_spread_pct",
]

def run_backtest(df, model):
    df = build_features(df)
    df = df.dropna(subset=FEATURES)

    # Score all rows at once — much faster than one at a time
    df["confidence"] = model.predict_proba(df[FEATURES])[:, 1]

    # Keep only rows that would trigger a trade signal
    signals = df[(df["gap_abs"] >= 0.5) & (df["confidence"] >= CONFIDENCE_THRESHOLD)].copy()
    print(f"  {len(signals):,} trade signals found", flush=True)

    trades = []

    for ticker in signals["ticker"].unique():
        tkr_all = df[df["ticker"] == ticker].copy()
        sig     = signals[signals["ticker"] == ticker]

        for _, row in sig.iterrows():
            entry_time    = row["timestamp"]
            entry_gap_pct = row["gap_pct"]
            entry_gap_abs = row["gap_abs"]
            spread        = row["bid_ask_spread_pct"] or 0.1
            confidence    = row["confidence"]
            direction     = "SHORT" if entry_gap_pct > 0 else "LONG"

            deadline = entry_time + timedelta(hours=HORIZON_HOURS)
            future   = tkr_all[(tkr_all["timestamp"] > entry_time) & (tkr_all["timestamp"] <= deadline)]

            if future.empty:
                continue

            closed_rows = future[future["gap_abs"] < CLOSE_THRESHOLD]
            if not closed_rows.empty:
                exit_row     = closed_rows.iloc[0]
                outcome      = "WIN"
            else:
                exit_row     = future.iloc[-1]
                outcome      = "LOSS"

            exit_time    = exit_row["timestamp"]
            exit_gap_abs = exit_row["gap_abs"]

            gap_captured = entry_gap_abs - exit_gap_abs
            spread_cost  = spread * POSITION_SIZE / 100
            pnl          = round((gap_captured / 100) * POSITION_SIZE - spread_cost, 4)
            hours_held   = round((exit_time - entry_time).total_seconds() / 3600, 1)

            trades.append({
                "ticker":        ticker,
                "entry_time":    entry_time.strftime("%Y-%m-%d %H:%M"),
                "exit_time":     exit_time.strftime("%Y-%m-%d %H:%M"),
                "direction":     direction,
                "entry_gap_pct": round(entry_gap_pct, 3),
                "exit_gap_pct":  round(exit_gap_abs if entry_gap_pct > 0 else -exit_gap_abs, 3),
                "spread_pct":    round(spread, 4),
                "confidence":    f"{confidence*100:.1f}%",
                "hours_held":    hours_held,
                "outcome":       outcome,
                "pnl_usd":       pnl,
            })

    return pd.DataFrame(trades)

def print_summary(trades_df):
    if trades_df.empty:
        print("No trades were generated. Try lowering CONFIDENCE_THRESHOLD.")
        return

    total      = len(trades_df)
    wins       = (trades_df["outcome"] == "WIN").sum()
    losses     = (trades_df["outcome"] == "LOSS").sum()
    win_rate   = wins / total * 100
    total_pnl  = trades_df["pnl_usd"].sum()
    avg_pnl    = trades_df["pnl_usd"].mean()
    best_trade = trades_df["pnl_usd"].max()
    worst_trade= trades_df["pnl_usd"].min()

    print("\n" + "="*50)
    print("  BACKTEST RESULTS")
    print("="*50)
    print(f"  Position size:   ${POSITION_SIZE:,} per trade")
    print(f"  Total trades:    {total}")
    print(f"  Wins:            {wins}  ({win_rate:.1f}%)")
    print(f"  Losses:          {losses}  ({100-win_rate:.1f}%)")
    print(f"  Total P&L:       ${total_pnl:+.2f}")
    print(f"  Avg P&L / trade: ${avg_pnl:+.2f}")
    print(f"  Best trade:      ${best_trade:+.2f}")
    print(f"  Worst trade:     ${worst_trade:+.2f}")
    print("="*50)
    print()

def save_excel(trades_df):
    OUTPUT = "backtest_results.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest Results"

    headers = list(trades_df.columns)

    # Styles
    navy        = "1E3A5F"
    white       = "FFFFFF"
    win_green   = "DEFFDE"
    loss_red    = "FFDEDE"
    alt_blue    = "F5F8FF"
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor=navy)
    header_font = Font(bold=True, color=white, size=10)

    for col_idx, col_name in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border

    for row_idx, (_, row) in enumerate(trades_df.iterrows(), 2):
        is_win = row["outcome"] == "WIN"
        for col_idx, col_name in enumerate(headers, 1):
            val            = row[col_name]
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(size=10)
            cell.border    = border
            if is_win:
                cell.fill = PatternFill("solid", fgColor=win_green)
            elif row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=alt_blue)

    col_widths = {
        "ticker": 8, "entry_time": 18, "exit_time": 18, "direction": 10,
        "entry_gap_pct": 14, "exit_gap_pct": 13, "spread_pct": 11,
        "confidence": 12, "hours_held": 12, "outcome": 10, "pnl_usd": 12,
    }
    for col_idx, col_name in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total     = len(trades_df)
    wins      = (trades_df["outcome"] == "WIN").sum()
    total_pnl = trades_df["pnl_usd"].sum()
    avg_pnl   = trades_df["pnl_usd"].mean()

    summary_rows = [
        ("Position size per trade", f"${POSITION_SIZE:,}"),
        ("Total trades",            total),
        ("Wins",                    wins),
        ("Losses",                  total - wins),
        ("Win rate",                f"{wins/total*100:.1f}%"),
        ("Total P&L",               f"${total_pnl:+.2f}"),
        ("Avg P&L per trade",       f"${avg_pnl:+.2f}"),
        ("Best trade",              f"${trades_df['pnl_usd'].max():+.2f}"),
        ("Worst trade",             f"${trades_df['pnl_usd'].min():+.2f}"),
    ]

    for r_idx, (label, value) in enumerate(summary_rows, 2):
        ws2.cell(row=r_idx, column=1, value=label).font  = Font(bold=True, size=11)
        ws2.cell(row=r_idx, column=2, value=value).font  = Font(size=11)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16

    wb.save(OUTPUT)
    print(f"  Full trade log saved to '{OUTPUT}'")
    print("  Open that file in Excel to see every trade.\n")

if __name__ == "__main__":
    print("Loading data...", flush=True)
    df = load_data()
    print(f"  {len(df):,} rows loaded", flush=True)

    print("Splitting into train/test by time...", flush=True)
    test_df = split_data(df)

    print("Loading model...", flush=True)
    with open(MODEL_FILE, "rb") as f:
        model = pickle.load(f)

    print("Running backtest on unseen test data only...", flush=True)
    trades_df = run_backtest(test_df, model)

    print_summary(trades_df)
    save_excel(trades_df)
