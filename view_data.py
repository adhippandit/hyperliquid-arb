import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

DB_FILE     = "gap_log.db"
OUTPUT_FILE = "gap_log_formatted.xlsx"

HEADER_BG    = "1E3A5F"   # dark navy
HEADER_FONT  = "FFFFFF"   # white
RED_BG       = "FFDEDE"   # light red  — stock overpriced on Hyperliquid
GREEN_BG     = "DEFFDE"   # light green — stock underpriced on Hyperliquid
ALT_ROW_BG   = "F5F8FF"   # very light blue for alternating rows

COL_WIDTHS = {
    "timestamp":          20,
    "ticker":              8,
    "real_price":         13,
    "hl_price":           13,
    "gap_usd":            11,
    "gap_pct":            10,
    "bid_ask_spread_pct": 18,
    "hours_to_open":      15,
    "market_open":        13,
    "day_of_week":        13,
    "hour_utc":           10,
    "vix":                 8,
}

def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def build_excel():
    if not os.path.exists(DB_FILE):
        print("No data file found. Run data_logger.py first to collect data.")
        return

    con = sqlite3.connect(DB_FILE)
    con.row_factory = sqlite3.Row
    raw  = con.execute("SELECT * FROM gaps ORDER BY id").fetchall()
    con.close()

    if not raw:
        print("No data yet in gap_log.db.")
        return

    rows    = [dict(r) for r in raw]
    headers = [k for k in rows[0].keys() if k != "id"]
    rows    = [{k: v for k, v in r.items() if k != "id"} for r in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Log"

    # --- Header row ---
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(bold=True, color=HEADER_FONT, size=10)
    border      = make_border()

    for col_idx, col_name in enumerate(headers, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        cell.border     = border

    # --- Data rows ---
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_BG)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(headers, start=1):
            raw   = row.get(col_name, "")
            # Convert numbers where possible
            try:
                value = float(raw)
            except (ValueError, TypeError):
                value = raw

            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(size=10)

            # Alternating row shading
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # --- Conditional formatting on gap_pct column ---
    gap_col_idx = headers.index("gap_pct") + 1
    gap_col_letter = get_column_letter(gap_col_idx)
    gap_range = f"{gap_col_letter}2:{gap_col_letter}{len(rows) + 1}"

    ws.conditional_formatting.add(gap_range, CellIsRule(
        operator="greaterThan", formula=["1"],
        fill=PatternFill("solid", fgColor="FFDEDE"),
        font=Font(bold=True, color="CC0000"),
    ))
    ws.conditional_formatting.add(gap_range, CellIsRule(
        operator="lessThan", formula=["-1"],
        fill=PatternFill("solid", fgColor="DEFFDE"),
        font=Font(bold=True, color="006600"),
    ))

    # --- Column widths ---
    for col_idx, col_name in enumerate(headers, start=1):
        width = COL_WIDTHS.get(col_name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto filter ---
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # --- Row height ---
    ws.row_dimensions[1].height = 20
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 16

    wb.save(OUTPUT_FILE)
    print(f"Done. Saved to '{OUTPUT_FILE}' — {len(rows)} rows.")
    print("Open that file in Excel.")

if __name__ == "__main__":
    build_excel()
